#!/usr/bin/env python3
"""
Damage the workbook on purpose and see whether the Summary cross-check notices.

This exists because of the lead-split post-mortem: a check that reads 0.00 on
good data proves nothing until you have watched it go non-zero on bad data.
It also records, honestly, what the check CANNOT see.
"""
import shutil
from datetime import date

import formulas
from openpyxl import load_workbook

import build_sheet as B

SCRATCH = "/tmp/claude-0/-root-workspace/2c7f0201-d8df-45a9-a9d4-79a3a914aa28/scratchpad"
SRC = f"{SCRATCH}/verify-ledger.xlsx"


def read(path, refs):
    sol = formulas.ExcelModel().loads(path).finish().calculate()
    fn = path.split("/")[-1]
    return {r: sol[f"'[{fn}]{s.upper()}'!{ref}"].value[0, 0] for r, (s, ref) in refs.items()}


def summary_rows():
    ws = load_workbook(SRC)["Summary"]
    out = {}
    for r in range(1, ws.max_row + 1):
        v = ws.cell(row=r, column=2).value
        if isinstance(v, str):
            out[v] = r
    return out


def main():
    rows = summary_rows()
    net_r = rows["NET FOR THE PERIOD"]
    out_r = rows["STILL OUTSTANDING"]
    chk = {k[:2]: v for k, v in rows.items() if k[:2] in ("1.", "2.", "3.", "4.")}
    assert len(chk) == 4, chk
    refs = {"net": ("Summary", f"C{net_r}"), "out": ("Summary", f"C{out_r}")}
    for k, v in chk.items():
        refs[k] = ("Summary", f"C{v}")

    base = read(SRC, refs)
    fired = lambda d: [k for k in ("1.", "2.", "3.", "4.") if abs(float(d[k])) > 1e-9]
    print(f"clean file   NET {float(base['net']):>10.2f}   outstanding "
          f"{float(base['out']):>10.2f}   checks firing: {fired(base) or 'none'}")
    assert not fired(base), "every check must read 0 on a clean file"

    # row 4 is the "10 leads Bob's own -> Olivia" fixture, so L4 = 80.00, not 0
    cases = [
        ("NET column typed over with a number", "Daily", "S5", 9999.0),
        ("'Olivia owes Bob' formula deleted",   "Daily", "L4", None),
        ("'Olivia owes Bob' typed over",        "Daily", "L4", 12.0),
        ("'Bob owes Olivia' overwritten",       "Daily", "M7", 0.0),
        ("other-cost adjustment typed over",    "Daily", "Q12", 500.0),
        ("price formula deleted for a day",     "Daily", "I4", None),
    ]
    # H (total leads) is display only — no money formula reads it, so damaging it
    # cannot move a settlement. Listed here so that is on the record, not a gap.
    harmless = [("total-leads column typed over", "Daily", "H4", 5)]
    caught = 0
    for label, sheet, ref, val in cases:
        p = f"{SCRATCH}/dmg.xlsx"
        shutil.copy(SRC, p)
        wb = load_workbook(p)
        wb[sheet][ref] = val
        wb.save(p)
        got = read(p, refs)
        f = fired(got)
        caught += bool(f)
        print(f"  {'CAUGHT ' if f else 'MISSED '} {label:<38} "
              f"fired by check {','.join(f) if f else '-':<10} "
              f"residual {max([float(got[k]) for k in f], key=abs) if f else 0:>10.2f}")

    print(f"\n  formula damage caught: {caught}/{len(cases)}   "
          f"(the residual equals the size of the error)")

    print("\n  damage that cannot move money, so nothing should fire:")
    for label, sheet, ref, val in harmless:
        p = f"{SCRATCH}/dmg3.xlsx"
        shutil.copy(SRC, p)
        wb = load_workbook(p); wb[sheet][ref] = val; wb.save(p)
        got = read(p, refs)
        moved = float(got["net"]) - float(base["net"])
        print(f"  {'ok    ' if not fired(got) and moved == 0 else 'PROBLEM'} {label:<38} "
              f"settlement moved {moved:>+12.2f}")

    print("\n  what the check CANNOT see (it reads 0.00, the money is still wrong):")
    blind = [
        ("a lead count typed wrong", "Daily", "B5", 999),
        ("leads put in the wrong lane", "Daily", "G8", 40),
        ("wrong person named as the payer", "Daily", "O8", "Olivia"),
        ("price on the Rates tab typed wrong", "Rates", "C6", 80.0),
    ]
    for label, sheet, ref, val in blind:
        p = f"{SCRATCH}/dmg2.xlsx"
        shutil.copy(SRC, p)
        wb = load_workbook(p)
        wb[sheet][ref] = val
        wb.save(p)
        got = read(p, refs)
        f = fired(got)
        moved = float(got["net"]) - float(base["net"])
        print(f"  {'silent' if not f else 'FIRED '} {label:<38} "
              f"but the settlement moved {moved:>+12.2f}")

    print("\n  => 0.00 means the formulas are intact. It does NOT mean the numbers are right.")


if __name__ == "__main__":
    main()

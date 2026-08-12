#!/usr/bin/env python3
"""
Build the Bob <-> Olivia lead ledger workbook.

Money map (the whole model, six lanes):

  source        destination        who bought the lead   who collects the sale   Bob <-> Olivia
  ------------------------------------------------------------------------------------------
  Bob's own  -> Olivia             Bob (ad spend)        Olivia buys it          Olivia owes Bob
  S          -> Olivia             Bob (pays S)          Olivia buys it          Olivia owes Bob
  GS         -> Olivia             Olivia (pays GS)      Olivia's own buyer      nothing
  Bob's own  -> Bob's buyer (GLU)  Bob (ad spend)        Bob (from GLU)          nothing*
  S          -> Bob's buyer (GLU)  Bob (pays S)          Bob (from GLU)          nothing*
  GS         -> Bob's buyer (GLU)  Olivia (pays GS)      Bob (from GLU)          Bob owes Olivia

  * plus a per-lead panel fee if one is agreed (rate defaults to 0)

NET (last column) is always signed as "Olivia pays Bob".
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import CellIsRule
from datetime import date

# ---------------------------------------------------------------- styling
INK = "1B2430"
HEAD_BG = "1B2430"
HEAD_FG = "FFFFFF"
BAND_BOB = "E8F0FB"      # lanes that end up with Olivia owing Bob
BAND_OLI = "FBEDE8"      # lanes that end up with Bob owing Olivia
BAND_NEU = "F1F3F5"      # neutral lanes
INPUT_BG = "FFF8E1"      # type here
CALC_BG = "F4F6F8"       # do not type here
NET_BG = "EAF6EC"

F_TITLE = Font(name="Calibri", size=16, bold=True, color=INK)
F_SUB = Font(name="Calibri", size=11, color="55606B")
F_HEAD = Font(name="Calibri", size=10, bold=True, color=HEAD_FG)
F_SECT = Font(name="Calibri", size=11, bold=True, color=INK)
F_BODY = Font(name="Calibri", size=11, color=INK)
F_BOLD = Font(name="Calibri", size=11, bold=True, color=INK)

THIN = Side(style="thin", color="C9D1D9")
BOX = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

MONEY = '#,##0.00'
COUNT = '#,##0'
DATEF = 'yyyy-mm-dd'

N_ROWS = 300          # daily rows pre-built with formulas
FIRST = 3             # first data row on Daily / Example
LAST = FIRST + N_ROWS - 1


def head(ws, row, values, widths=None, height=34):
    for i, v in enumerate(values, start=1):
        c = ws.cell(row=row, column=i, value=v)
        c.font = F_HEAD
        c.fill = PatternFill("solid", fgColor=HEAD_BG)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = BOX
    ws.row_dimensions[row].height = height
    if widths:
        for i, w in enumerate(widths, start=1):
            ws.column_dimensions[get_column_letter(i)].width = w


def title(ws, text, sub=None):
    ws["A1"] = text
    ws["A1"].font = F_TITLE
    ws.row_dimensions[1].height = 24
    if sub:
        ws["A2"] = sub
        ws["A2"].font = F_SUB


# ================================================================ 1. HOW IT WORKS
def sheet_how(wb):
    ws = wb.create_sheet("How it works")
    ws.sheet_view.showGridLines = False
    title(ws, "How this ledger works",
          "Read this once. Then you only ever type on the Daily tab.")

    ws.column_dimensions["A"].width = 3
    for col, w in zip("BCDEFG", [26, 22, 22, 24, 24, 34]):
        ws.column_dimensions[col].width = w

    r = 4
    ws.cell(row=r, column=2, value="THE SIX LANES — every lead is one of these").font = F_SECT
    r += 1
    head_vals = ["Lead came from", "Sold to", "Who paid for the lead",
                 "Who collects the sale", "Effect between Bob and Olivia"]
    for i, v in enumerate(head_vals):
        c = ws.cell(row=r, column=2 + i, value=v)
        c.font = F_HEAD
        c.fill = PatternFill("solid", fgColor=HEAD_BG)
        c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        c.border = BOX
    ws.row_dimensions[r].height = 30

    lanes = [
        ("Bob's own campaign", "Olivia", "Bob (his ad spend)", "Olivia buys it",
         "OLIVIA OWES BOB — price per lead", BAND_BOB),
        ("S", "Olivia", "Bob (pays S)", "Olivia buys it",
         "OLIVIA OWES BOB — same price per lead", BAND_BOB),
        ("GS", "Olivia", "Olivia (pays GS)", "Olivia's own buyer",
         "nothing — both sides are Olivia's", BAND_NEU),
        ("Bob's own campaign", "Bob's private buyer (GLU)", "Bob (his ad spend)", "Bob (from GLU)",
         "nothing — both sides are Bob's *", BAND_NEU),
        ("S", "Bob's private buyer (GLU)", "Bob (pays S)", "Bob (from GLU)",
         "nothing — both sides are Bob's *", BAND_NEU),
        ("GS", "Bob's private buyer (GLU)", "Olivia (pays GS)", "Bob (from GLU)",
         "BOB OWES OLIVIA — transfer price per lead", BAND_OLI),
    ]
    for lane in lanes:
        r += 1
        for i, v in enumerate(lane[:5]):
            c = ws.cell(row=r, column=2 + i, value=v)
            c.font = F_BODY
            c.fill = PatternFill("solid", fgColor=lane[5])
            c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
            c.border = BOX
        ws.row_dimensions[r].height = 28

    r += 2
    ws.cell(row=r, column=2,
            value="* plus a per-lead panel fee, if you agree one. The rate is 0 by default, "
                  "so it does nothing until you put a number in it.").font = F_SUB
    r += 3

    ws.cell(row=r, column=2, value="THE ONE RULE").font = F_SECT
    r += 1
    for line in [
        "NET  =  (what Olivia owes Bob)  -  (what Bob owes Olivia)  +/-  the other cost, if you settle it",
        "",
        "The NET is always written from Olivia's side. Positive means Olivia pays Bob. "
        "Negative means Bob pays Olivia. It is the last column on the Daily tab.",
    ]:
        ws.cell(row=r, column=2, value=line).font = F_BOLD if line.startswith("NET") else F_BODY
        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=6)
        ws.cell(row=r, column=2).alignment = Alignment(wrap_text=True, vertical="top")
        ws.row_dimensions[r].height = 30 if len(line) > 90 else 18
        r += 1

    r += 2
    ws.cell(row=r, column=2, value="WHAT TO DO EACH DAY").font = F_SECT
    r += 1
    steps = [
        "1.  Open the Daily tab. Put the date in column A.",
        "2.  Put the six lead counts in columns B to G. Leave a lane blank if it did not happen.",
        "3.  The three price columns fill themselves from the Rates tab. Type over one only if the price changed that day.",
        "4.  If there was any other cost, put the amount, who paid it, and YES or NO in columns N, O, P.",
        "5.  Read the last column. It tells you who pays whom, and how much.",
    ]
    for s in steps:
        ws.cell(row=r, column=2, value=s).font = F_BODY
        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=6)
        ws.cell(row=r, column=2).alignment = Alignment(wrap_text=True, vertical="top")
        r += 1

    r += 2
    ws.cell(row=r, column=2, value="COLOUR CODE").font = F_SECT
    r += 1
    for label, colr in [("Yellow cells — you type in these", INPUT_BG),
                        ("Grey cells — worked out for you, do not type in these", CALC_BG),
                        ("Green cell — the answer for that day", NET_BG)]:
        c = ws.cell(row=r, column=2, value=label)
        c.font = F_BODY
        c.fill = PatternFill("solid", fgColor=colr)
        c.border = BOX
        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=4)
        r += 1

    r += 2
    ws.cell(row=r, column=2, value="THINGS THIS SHEET ASSUMES — tell me if any is wrong").font = F_SECT
    r += 1
    for s in [
        "a.  Olivia pays Bob the same price per lead whether the lead came from Bob's campaign or from S.",
        "b.  A GS lead that ends up with Bob's private buyer is paid for by Bob at the transfer price on the Rates tab.",
        "c.  Other cost set to YES means the other party pays the payer back in full, not half.",
        "d.  All of Bob's private buyers are counted together. Only GLU is named so far.",
    ]:
        ws.cell(row=r, column=2, value=s).font = F_BODY
        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=6)
        ws.cell(row=r, column=2).alignment = Alignment(wrap_text=True, vertical="top")
        r += 1

    return ws


# ================================================================ 2. RATES
def sheet_rates(wb):
    ws = wb.create_sheet("Rates")
    ws.sheet_view.showGridLines = False
    title(ws, "Rates — set these once",
          "The Daily tab reads the three settlement rates from here.")

    ws.column_dimensions["A"].width = 3
    ws.column_dimensions["B"].width = 56
    ws.column_dimensions["C"].width = 14
    ws.column_dimensions["D"].width = 22
    ws.column_dimensions["E"].width = 52

    r = 4
    ws.cell(row=r, column=2,
            value="SETTLEMENT RATES — these three are the only ones that move money "
                  "between Bob and Olivia").font = F_SECT
    r += 1
    for i, v in enumerate(["Rate", "Value", "Direction", "Notes"]):
        c = ws.cell(row=r, column=2 + i, value=v)
        c.font = F_HEAD
        c.fill = PatternFill("solid", fgColor=HEAD_BG)
        c.alignment = Alignment(horizontal="left", vertical="center")
        c.border = BOX
    ws.row_dimensions[r].height = 22

    # rows 6, 7, 8 -> C6, C7, C8 are the live rates
    rows = [
        ("Bob to Olivia — price per lead", 8.00, "Olivia pays Bob",
         "Same price whether the lead came from Bob's campaign or from S."),
        ("GS lead sold to Bob's private buyer — price per lead", "=C15", "Bob pays Olivia",
         "Defaults to what GS costs Olivia. Type a number over it if you agreed a different price."),
        ("Panel fee per lead on Bob's private-buyer leads", 0.00, "Bob pays Olivia",
         "Leave at 0 if Bob pays nothing for using Olivia's panel."),
    ]
    rate_rows = {}
    for name, val, direction, note in rows:
        r += 1
        rate_rows[name] = r
        b = ws.cell(row=r, column=2, value=name); b.font = F_BODY; b.border = BOX
        c = ws.cell(row=r, column=3, value=val)
        c.font = F_BOLD; c.border = BOX; c.number_format = MONEY
        c.fill = PatternFill("solid", fgColor=INPUT_BG)
        d = ws.cell(row=r, column=4, value=direction); d.font = F_BODY; d.border = BOX
        e = ws.cell(row=r, column=5, value=note); e.font = F_SUB; e.border = BOX
        e.alignment = Alignment(wrap_text=True, vertical="center")
        ws.row_dimensions[r].height = 30

    r += 3
    ws.cell(row=r, column=2,
            value="REFERENCE ONLY — recorded so the prices are written down. "
                  "Nothing here changes what Bob and Olivia owe each other.").font = F_SECT
    r += 1
    for i, v in enumerate(["Rate", "Value", "Who pays it", "Notes"]):
        c = ws.cell(row=r, column=2 + i, value=v)
        c.font = F_HEAD
        c.fill = PatternFill("solid", fgColor="55606B")
        c.alignment = Alignment(horizontal="left", vertical="center")
        c.border = BOX

    ref = [
        ("Bob's own campaign — cost per lead", 0.00, "Bob", "Bob's ad spend divided by leads."),
        ("S — cost per lead", 0.00, "Bob", "What Bob pays S."),
        ("GS — cost per lead", 6.00, "Olivia", "What Olivia pays GS. Row 7 above defaults to this."),
        ("Bob to GLU — price per lead", 0.00, "GLU pays Bob", "Bob's private-buyer price."),
    ]
    for name, val, who, note in ref:
        r += 1
        b = ws.cell(row=r, column=2, value=name); b.font = F_BODY; b.border = BOX
        c = ws.cell(row=r, column=3, value=val)
        c.font = F_BODY; c.border = BOX; c.number_format = MONEY
        c.fill = PatternFill("solid", fgColor=INPUT_BG)
        d = ws.cell(row=r, column=4, value=who); d.font = F_BODY; d.border = BOX
        e = ws.cell(row=r, column=5, value=note); e.font = F_SUB; e.border = BOX

    r += 3
    ws.cell(row=r, column=2, value="Currency").font = F_BOLD
    cc = ws.cell(row=r, column=3, value="USD")
    cc.font = F_BODY; cc.border = BOX
    cc.fill = PatternFill("solid", fgColor=INPUT_BG)
    ws.cell(row=r, column=5, value="Label only. Every amount in the file is in this currency.").font = F_SUB

    return ws


# ================================================================ 3. DAILY / EXAMPLE
DAILY_HEADERS = [
    "Date",
    "Leads\nBob's own\n-> Olivia",
    "Leads\nS\n-> Olivia",
    "Leads\nGS\n-> Olivia",
    "Leads\nBob's own\n-> GLU",
    "Leads\nS\n-> GLU",
    "Leads\nGS\n-> GLU",
    "Total\nleads",
    "Price\nBob -> Olivia\nper lead",
    "Price\nGS -> GLU\nper lead",
    "Panel fee\nper lead",
    "OLIVIA\nOWES BOB",
    "BOB\nOWES OLIVIA",
    "Other cost\namount",
    "Other cost\npaid by",
    "Settle it?\nYES / NO",
    "Other cost\nadjustment",
    "Who pays whom",
    "NET\n(+ Olivia pays Bob)\n(- Bob pays Olivia)",
]
DAILY_WIDTHS = [12, 11, 10, 10, 11, 10, 10, 8, 12, 12, 10, 13, 13, 11, 11, 11, 12, 30, 18]


def daily_formulas(r, rates=True):
    """Formulas for one data row. rates=False -> Example tab uses fixed prices."""
    p_bo = "Rates!$C$6" if rates else None
    p_gs = "Rates!$C$7" if rates else None
    p_pf = "Rates!$C$8" if rates else None
    g = f'IF($A{r}="","",'          # blank date -> blank row
    f = {}
    f["H"] = f'={g}SUM($B{r}:$G{r}))'
    if rates:
        f["I"] = f'={g}{p_bo})'
        f["J"] = f'={g}{p_gs})'
        f["K"] = f'={g}{p_pf})'
    f["L"] = f'={g}($B{r}+$C{r})*$I{r})'
    f["M"] = f'={g}$G{r}*$J{r}+($E{r}+$F{r}+$G{r})*$K{r})'
    f["Q"] = f'={g}IF($P{r}="YES",IF($O{r}="Bob",$N{r},-$N{r}),0))'
    f["S"] = f'={g}$L{r}-$M{r}+$Q{r})'
    f["R"] = (f'={g}IF(ROUND($S{r},2)>0,"Olivia pays Bob "&TEXT($S{r},"#,##0.00"),'
              f'IF(ROUND($S{r},2)<0,"Bob pays Olivia "&TEXT(-$S{r},"#,##0.00"),"Nothing owed")))')
    return f


def build_daily(ws, rates=True, example_rows=None):
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "B3"

    if rates:
        ws["A1"] = "Daily ledger — type the date and the six lead counts, read the last column"
    else:
        ws["A1"] = "Example — three worked days, numbers typed in so you can check the maths"
    ws["A1"].font = F_TITLE
    ws.row_dimensions[1].height = 22

    head(ws, 2, DAILY_HEADERS, DAILY_WIDTHS, height=46)

    lane_fill = {
        "B": BAND_BOB, "C": BAND_BOB, "D": BAND_NEU,
        "E": BAND_NEU, "F": BAND_NEU, "G": BAND_OLI,
    }
    for col, colr in lane_fill.items():
        c = ws[f"{col}2"]
        c.fill = PatternFill("solid", fgColor=colr)
        c.font = Font(name="Calibri", size=10, bold=True, color=INK)

    n = len(example_rows) if example_rows else N_ROWS
    for i in range(n):
        r = FIRST + i
        ws.row_dimensions[r].height = 17

        a = ws.cell(row=r, column=1)
        a.number_format = DATEF
        a.fill = PatternFill("solid", fgColor=INPUT_BG)
        a.border = BOX
        a.font = F_BODY

        for col in "BCDEFG":
            c = ws[f"{col}{r}"]
            c.number_format = COUNT
            c.fill = PatternFill("solid", fgColor=INPUT_BG)
            c.border = BOX
            c.font = F_BODY

        f = daily_formulas(r, rates=rates)

        c = ws[f"H{r}"]; c.value = f["H"]; c.number_format = COUNT
        c.fill = PatternFill("solid", fgColor=CALC_BG); c.border = BOX; c.font = F_BODY

        for col in "IJK":
            c = ws[f"{col}{r}"]
            if rates:
                c.value = f[col]
                c.fill = PatternFill("solid", fgColor=CALC_BG)
            else:
                c.fill = PatternFill("solid", fgColor=INPUT_BG)
            c.number_format = MONEY
            c.border = BOX
            c.font = F_BODY

        for col in "LM":
            c = ws[f"{col}{r}"]; c.value = f[col]; c.number_format = MONEY
            c.fill = PatternFill("solid", fgColor=CALC_BG); c.border = BOX; c.font = F_BODY

        for col in "NOP":
            c = ws[f"{col}{r}"]
            c.fill = PatternFill("solid", fgColor=INPUT_BG); c.border = BOX; c.font = F_BODY
        ws[f"N{r}"].number_format = MONEY

        c = ws[f"Q{r}"]; c.value = f["Q"]; c.number_format = MONEY
        c.fill = PatternFill("solid", fgColor=CALC_BG); c.border = BOX; c.font = F_BODY

        c = ws[f"R{r}"]; c.value = f["R"]
        c.fill = PatternFill("solid", fgColor=CALC_BG); c.border = BOX; c.font = F_BODY
        c.alignment = Alignment(horizontal="left")

        c = ws[f"S{r}"]; c.value = f["S"]; c.number_format = MONEY
        c.fill = PatternFill("solid", fgColor=NET_BG); c.border = BOX
        c.font = Font(name="Calibri", size=11, bold=True, color=INK)

    # dropdowns
    dv_who = DataValidation(type="list", formula1='"Bob,Olivia"', allow_blank=True,
                            showDropDown=False)
    dv_yn = DataValidation(type="list", formula1='"YES,NO"', allow_blank=True,
                           showDropDown=False)
    dv_who.error = "Type Bob or Olivia."
    dv_yn.error = "Type YES or NO."
    ws.add_data_validation(dv_who)
    ws.add_data_validation(dv_yn)
    dv_who.add(f"O{FIRST}:O{FIRST + n - 1}")
    dv_yn.add(f"P{FIRST}:P{FIRST + n - 1}")

    # colour the net
    rng = f"S{FIRST}:S{FIRST + n - 1}"
    ws.conditional_formatting.add(rng, CellIsRule(
        operator="greaterThan", formula=["0"],
        fill=PatternFill("solid", bgColor="C6EFCE"), font=Font(color="1B5E20", bold=True)))
    ws.conditional_formatting.add(rng, CellIsRule(
        operator="lessThan", formula=["0"],
        fill=PatternFill("solid", bgColor="FFC7CE"), font=Font(color="9C1F1F", bold=True)))

    # example data
    if example_rows:
        for i, row in enumerate(example_rows):
            r = FIRST + i
            ws.cell(row=r, column=1, value=row["date"])
            for j, col in enumerate("BCDEFG"):
                ws[f"{col}{r}"] = row["counts"][j]
            ws[f"I{r}"] = row["p_bo"]
            ws[f"J{r}"] = row["p_gs"]
            ws[f"K{r}"] = row["p_pf"]
            if row["other"] is not None:
                ws[f"N{r}"] = row["other"][0]
                ws[f"O{r}"] = row["other"][1]
                ws[f"P{r}"] = row["other"][2]

    ws.auto_filter.ref = f"A2:S{FIRST + n - 1}"
    return ws


# ================================================================ 4. SUMMARY
def sheet_summary(wb):
    ws = wb.create_sheet("Summary")
    ws.sheet_view.showGridLines = False
    title(ws, "Summary — the balance between Bob and Olivia",
          "Change the two dates to settle a week, a month, or everything.")

    ws.column_dimensions["A"].width = 3
    ws.column_dimensions["B"].width = 52
    ws.column_dimensions["C"].width = 16
    ws.column_dimensions["D"].width = 16
    ws.column_dimensions["E"].width = 44

    D = f"Daily!$A${FIRST}:$A${LAST}"

    def sumifs(col):
        return (f'SUMIFS(Daily!${col}${FIRST}:${col}${LAST},'
                f'{D},">="&$C$4,{D},"<="&$C$5)')

    r = 4
    ws.cell(row=r, column=2, value="From date").font = F_BOLD
    c = ws.cell(row=r, column=3, value=date(2000, 1, 1))
    c.number_format = DATEF; c.border = BOX
    c.fill = PatternFill("solid", fgColor=INPUT_BG)
    r += 1
    ws.cell(row=r, column=2, value="To date").font = F_BOLD
    c = ws.cell(row=r, column=3, value=date(2099, 12, 31))
    c.number_format = DATEF; c.border = BOX
    c.fill = PatternFill("solid", fgColor=INPUT_BG)

    r += 2
    ws.cell(row=r, column=2, value="LEADS IN THE PERIOD").font = F_SECT
    r += 1
    for i, v in enumerate(["Lane", "Leads"]):
        c = ws.cell(row=r, column=2 + i, value=v)
        c.font = F_HEAD; c.fill = PatternFill("solid", fgColor=HEAD_BG); c.border = BOX

    lanes = [
        ("Bob's own campaign -> Olivia", "B", BAND_BOB),
        ("S -> Olivia", "C", BAND_BOB),
        ("GS -> Olivia", "D", BAND_NEU),
        ("Bob's own campaign -> Bob's private buyer", "E", BAND_NEU),
        ("S -> Bob's private buyer", "F", BAND_NEU),
        ("GS -> Bob's private buyer", "G", BAND_OLI),
    ]
    first_lane = r + 1
    for name, col, colr in lanes:
        r += 1
        b = ws.cell(row=r, column=2, value=name); b.font = F_BODY; b.border = BOX
        b.fill = PatternFill("solid", fgColor=colr)
        c = ws.cell(row=r, column=3, value=f"={sumifs(col)}")
        c.number_format = COUNT; c.border = BOX; c.font = F_BODY
        c.fill = PatternFill("solid", fgColor=CALC_BG)
    r += 1
    ws.cell(row=r, column=2, value="Total leads").font = F_BOLD
    c = ws.cell(row=r, column=3, value=f"=SUM(C{first_lane}:C{r - 1})")
    c.number_format = COUNT; c.border = BOX; c.font = F_BOLD
    c.fill = PatternFill("solid", fgColor=CALC_BG)

    r += 2
    ws.cell(row=r, column=2, value="MONEY IN THE PERIOD").font = F_SECT
    money_rows = {}
    for label, col, note in [
        ("Olivia owes Bob — lead sales", "L",
         "Every lead Bob sold her, from his own campaign or from S."),
        ("Bob owes Olivia — GS leads to his buyers, and panel fee", "M",
         "GS leads Olivia paid for that went to Bob's private buyer."),
        ("Other cost adjustment", "Q",
         "Plus if Bob paid it, minus if Olivia paid it. Only the rows marked YES."),
    ]:
        r += 1
        b = ws.cell(row=r, column=2, value=label); b.font = F_BODY; b.border = BOX
        c = ws.cell(row=r, column=3, value=f"={sumifs(col)}")
        c.number_format = MONEY; c.border = BOX; c.font = F_BODY
        c.fill = PatternFill("solid", fgColor=CALC_BG)
        e = ws.cell(row=r, column=5, value=note); e.font = F_SUB
        money_rows[col] = r

    r += 1
    b = ws.cell(row=r, column=2, value="NET FOR THE PERIOD"); b.font = F_BOLD; b.border = BOX
    c = ws.cell(row=r, column=3,
                value=f"=C{money_rows['L']}-C{money_rows['M']}+C{money_rows['Q']}")
    c.number_format = MONEY; c.border = BOX
    c.font = Font(name="Calibri", size=12, bold=True, color=INK)
    c.fill = PatternFill("solid", fgColor=NET_BG)
    net_row = r

    # ---- checks, each computed by a route that does NOT read the column it checks
    def dcol(col):
        return f"Daily!${col}${FIRST}:${col}${LAST}"

    DA = dcol("A")
    W = f'({DA}>=$C$4)*({DA}<=$C$5)'
    Lr, Mr, Qr = money_rows["L"], money_rows["M"], money_rows["Q"]

    r += 2
    ws.cell(row=r, column=2, value="CHECKS — all four must read 0").font = F_SECT
    ws.cell(row=r, column=5,
            value="Each one rebuilds a figure a different way. A number other than 0 means a "
                  "formula has been typed over.").font = F_SUB

    checks = [
        ("1. Lead money, rebuilt from the counts and the prices",
         f'=ROUND(SUMPRODUCT({W}*({dcol("B")}+{dcol("C")}),{dcol("I")})'
         f'-SUMPRODUCT({W}*{dcol("G")},{dcol("J")})'
         f'-SUMPRODUCT({W}*({dcol("E")}+{dcol("F")}+{dcol("G")}),{dcol("K")})'
         f'-(C{Lr}-C{Mr}),6)',
         "Multiplies the lead counts by the prices itself, instead of trusting "
         "the two money columns."),
        ("2. Other cost, rebuilt from the amount, payer and YES/NO",
         f'=ROUND(SUMIFS({dcol("N")},{dcol("P")},"YES",{dcol("O")},"Bob",'
         f'{DA},">="&$C$4,{DA},"<="&$C$5)'
         f'-SUMIFS({dcol("N")},{dcol("P")},"YES",{dcol("O")},"Olivia",'
         f'{DA},">="&$C$4,{DA},"<="&$C$5)-C{Qr},6)',
         "Re-reads the three other-cost columns instead of trusting the adjustment column."),
        ("3. The daily NET column adds up to the same net",
         f'=ROUND(C{Lr}-C{Mr}+C{Qr}-{sumifs("S")},6)',
         "Catches a NET cell that has had a number typed straight into it."),
        ("4. Days in the period with no price on them",
         f'=SUMPRODUCT({W}*({dcol("I")}=""))',
         "A day priced at nothing earns nothing. Should never be anything but 0."),
    ]
    for label, formula, note in checks:
        r += 1
        b = ws.cell(row=r, column=2, value=label); b.font = F_SUB; b.border = BOX
        c = ws.cell(row=r, column=3, value=formula)
        c.number_format = MONEY; c.border = BOX; c.font = F_SUB
        c.fill = PatternFill("solid", fgColor=CALC_BG)
        ws.conditional_formatting.add(f"C{r}", CellIsRule(
            operator="notEqual", formula=["0"],
            fill=PatternFill("solid", bgColor="FFC7CE"), font=Font(color="9C1F1F", bold=True)))
        ws.cell(row=r, column=5, value=note).font = F_SUB

    r += 1
    ws.cell(row=r, column=2,
            value="0 means the formulas are intact. It does NOT mean the numbers are right — "
                  "a lead count typed wrong, or a lead logged in the wrong lane, passes all "
                  "four silently.").font = F_SUB
    ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=5)
    ws.cell(row=r, column=2).alignment = Alignment(wrap_text=True, vertical="top")
    ws.row_dimensions[r].height = 28

    r += 3
    ws.cell(row=r, column=2, value="MONEY ALREADY TRANSFERRED").font = F_SECT
    ws.cell(row=r, column=5,
            value="Log every payment here, or the balance below will keep asking for money "
                  "that has already moved.").font = F_SUB
    r += 1
    pay_head = r
    for i, v in enumerate(["Date", "Paid by", "Amount", "Note"]):
        c = ws.cell(row=r, column=2 + i, value=v)
        c.font = F_HEAD; c.fill = PatternFill("solid", fgColor=HEAD_BG); c.border = BOX
    pay_first = r + 1
    pay_last = pay_first + 24
    for rr in range(pay_first, pay_last + 1):
        for cc in range(2, 6):
            c = ws.cell(row=rr, column=cc)
            c.fill = PatternFill("solid", fgColor=INPUT_BG); c.border = BOX; c.font = F_BODY
        ws.cell(row=rr, column=2).number_format = DATEF
        ws.cell(row=rr, column=4).number_format = MONEY

    dv_who = DataValidation(type="list", formula1='"Bob,Olivia"', allow_blank=True,
                            showDropDown=False)
    ws.add_data_validation(dv_who)
    dv_who.add(f"C{pay_first}:C{pay_last}")

    r = pay_last + 2
    ws.cell(row=r, column=2, value="Paid by Olivia to Bob").font = F_BODY
    c = ws.cell(row=r, column=3,
                value=f'=SUMIFS($D${pay_first}:$D${pay_last},$C${pay_first}:$C${pay_last},"Olivia",'
                      f'$B${pay_first}:$B${pay_last},">="&$C$4,'
                      f'$B${pay_first}:$B${pay_last},"<="&$C$5)')
    c.number_format = MONEY; c.border = BOX; c.font = F_BODY
    c.fill = PatternFill("solid", fgColor=CALC_BG)
    paid_o = r
    r += 1
    ws.cell(row=r, column=2, value="Paid by Bob to Olivia").font = F_BODY
    c = ws.cell(row=r, column=3,
                value=f'=SUMIFS($D${pay_first}:$D${pay_last},$C${pay_first}:$C${pay_last},"Bob",'
                      f'$B${pay_first}:$B${pay_last},">="&$C$4,'
                      f'$B${pay_first}:$B${pay_last},"<="&$C$5)')
    c.number_format = MONEY; c.border = BOX; c.font = F_BODY
    c.fill = PatternFill("solid", fgColor=CALC_BG)
    paid_b = r

    r += 2
    b = ws.cell(row=r, column=2, value="STILL OUTSTANDING"); b.font = F_BOLD; b.border = BOX
    c = ws.cell(row=r, column=3, value=f"=C{net_row}-C{paid_o}+C{paid_b}")
    c.number_format = MONEY; c.border = BOX
    c.font = Font(name="Calibri", size=12, bold=True, color=INK)
    c.fill = PatternFill("solid", fgColor=NET_BG)
    out_row = r
    ws.conditional_formatting.add(f"C{r}", CellIsRule(
        operator="greaterThan", formula=["0"],
        fill=PatternFill("solid", bgColor="C6EFCE"), font=Font(color="1B5E20", bold=True)))
    ws.conditional_formatting.add(f"C{r}", CellIsRule(
        operator="lessThan", formula=["0"],
        fill=PatternFill("solid", bgColor="FFC7CE"), font=Font(color="9C1F1F", bold=True)))

    r += 2
    c = ws.cell(row=r, column=2,
                value=f'=IF(ROUND(C{out_row},2)>0,"OLIVIA PAYS BOB "&TEXT(C{out_row},"#,##0.00"),'
                      f'IF(ROUND(C{out_row},2)<0,"BOB PAYS OLIVIA "&TEXT(-C{out_row},"#,##0.00"),'
                      f'"NOTHING OWED — SETTLED"))')
    c.font = Font(name="Calibri", size=14, bold=True, color=INK)
    ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=4)
    ws.row_dimensions[r].height = 26

    return ws


# ================================================================ build
def main():
    wb = Workbook()
    wb.remove(wb.active)

    sheet_how(wb)
    sheet_rates(wb)
    build_daily(wb.create_sheet("Daily"), rates=True)
    sheet_summary(wb)

    example = [
        dict(date=date(2026, 8, 10), counts=[40, 0, 25, 10, 0, 12],
             p_bo=8.00, p_gs=6.00, p_pf=0.00, other=None),
        dict(date=date(2026, 8, 11), counts=[0, 35, 18, 0, 8, 20],
             p_bo=8.00, p_gs=6.00, p_pf=0.00, other=(30.00, "Bob", "YES")),
        dict(date=date(2026, 8, 12), counts=[5, 0, 30, 2, 0, 45],
             p_bo=8.00, p_gs=6.00, p_pf=0.00, other=(50.00, "Olivia", "YES")),
    ]
    build_daily(wb.create_sheet("Example"), rates=False, example_rows=example)

    out = "/root/workspace/lead-ledger-bob-olivia/lead-ledger-bob-olivia.xlsx"
    wb.save(out)
    print("wrote", out)


if __name__ == "__main__":
    main()

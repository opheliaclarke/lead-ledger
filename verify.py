#!/usr/bin/env python3
"""
Independent verification of lead-ledger-bob-olivia.xlsx.

Two things are checked, and they are computed by different routes:

  MODEL   an exact-Fraction implementation of the six-lane money map, written
          from the specification, not from the spreadsheet.
  SHEET   the workbook's own formulas, evaluated by the `formulas` package
          (a real Excel formula engine), on a filled copy of the file.

They must agree cell for cell. Anything else is a bug in the workbook.
"""
import random
import sys
from datetime import date, datetime
from fractions import Fraction as Fr

import formulas
from openpyxl import load_workbook

import build_sheet as B

TEST_ROWS = 30
XL = "/tmp/claude-0/-root-workspace/2c7f0201-d8df-45a9-a9d4-79a3a914aa28/scratchpad/verify-ledger.xlsx"

FAILS = []


def to_fr(x):
    return Fr(float(x))


def check(name, got, want, tol=Fr(1, 10 ** 9)):
    if isinstance(want, str):
        ok = str(got) == want
    else:
        ok = abs(to_fr(got) - to_fr(want)) <= tol
    if not ok:
        FAILS.append(f"{name}: got {got!r}, want {want!r}")
    return ok


# ------------------------------------------------------------------ MODEL
def model_row(counts, p_bo, p_gs, p_pf, other):
    """counts = [bob->O, S->O, GS->O, bob->GLU, S->GLU, GS->GLU]"""
    b, c, d, e, f, g = (Fr(x) for x in counts)
    p_bo, p_gs, p_pf = Fr(str(p_bo)), Fr(str(p_gs)), Fr(str(p_pf))

    olivia_owes_bob = (b + c) * p_bo                     # lanes 1 and 2
    bob_owes_olivia = g * p_gs + (e + f + g) * p_pf      # lane 6, plus panel fee
    # lanes 3, 4, 5 create nothing between them, by construction: d never appears.

    adj = Fr(0)
    if other is not None:
        amt, payer, settle = other
        if settle == "YES":
            adj = Fr(str(amt)) if payer == "Bob" else -Fr(str(amt))

    net = olivia_owes_bob - bob_owes_olivia + adj
    total = b + c + d + e + f + g
    return dict(H=total, L=olivia_owes_bob, M=bob_owes_olivia, Q=adj, S=net)


def direction_text(net):
    r = round(float(net), 2)
    if r > 0:
        return f"Olivia pays Bob {float(net):,.2f}"
    if r < 0:
        return f"Bob pays Olivia {float(-net):,.2f}"
    return "Nothing owed"


# ------------------------------------------------------------------ fixtures
def make_rows(n, rng):
    rows = []
    # deliberate edge cases first
    fixed = [
        ([0, 0, 0, 0, 0, 0], None),                        # a day with nothing
        ([10, 0, 0, 0, 0, 0], None),                       # only Bob's own to Olivia
        ([0, 0, 99, 0, 0, 0], None),                       # only GS to Olivia -> must be 0
        ([0, 0, 0, 7, 4, 0], None),                        # only Bob's leads to his own buyer -> 0
        ([0, 0, 0, 0, 0, 25], None),                       # only GS to GLU -> Bob owes
        ([3, 2, 1, 1, 1, 1], (40.0, "Bob", "YES")),        # cost Bob paid, settled
        ([3, 2, 1, 1, 1, 1], (40.0, "Olivia", "YES")),     # cost Olivia paid, settled
        ([3, 2, 1, 1, 1, 1], (40.0, "Bob", "NO")),         # cost logged, not settled
        ([3, 2, 1, 1, 1, 1], (40.0, "Olivia", "NO")),
        ([0, 0, 0, 0, 0, 0], (12.5, "Olivia", "YES")),     # cost only, no leads
    ]
    for counts, other in fixed:
        rows.append(dict(counts=counts, other=other))
    while len(rows) < n:
        counts = [rng.choice([0, 0, rng.randint(1, 120)]) for _ in range(6)]
        other = None
        if rng.random() < 0.4:
            other = (round(rng.uniform(5, 300), 2),
                     rng.choice(["Bob", "Olivia"]),
                     rng.choice(["YES", "NO"]))
        rows.append(dict(counts=counts, other=other))
    return rows[:n]


def fill(rows, p_bo, p_gs, p_pf, payments):
    """Build a small copy of the real workbook and type the test data into it."""
    B.N_ROWS = TEST_ROWS
    B.LAST = B.FIRST + TEST_ROWS - 1
    from openpyxl import Workbook
    wb = Workbook()
    wb.remove(wb.active)
    B.sheet_how(wb)
    B.sheet_rates(wb)
    B.build_daily(wb.create_sheet("Daily"), rates=True)
    B.sheet_summary(wb)

    ws = wb["Rates"]
    ws["C6"] = p_bo
    ws["C7"] = p_gs          # overrides the =C15 default, on purpose
    ws["C8"] = p_pf

    d = wb["Daily"]
    start = date(2026, 8, 1)
    for i, row in enumerate(rows):
        r = B.FIRST + i
        d.cell(row=r, column=1, value=date.fromordinal(start.toordinal() + i))
        for j, col in enumerate("BCDEFG"):
            d[f"{col}{r}"] = row["counts"][j]
        if row["other"]:
            d[f"N{r}"], d[f"O{r}"], d[f"P{r}"] = row["other"]

    s = wb["Summary"]
    pay_first = None
    for r in range(1, s.max_row + 1):
        if s.cell(row=r, column=2).value == "Date" and s.cell(row=r, column=3).value == "Paid by":
            pay_first = r + 1
            break
    for i, (pd, who, amt) in enumerate(payments):
        s.cell(row=pay_first + i, column=2, value=pd)
        s.cell(row=pay_first + i, column=3, value=who)
        s.cell(row=pay_first + i, column=4, value=amt)

    wb.save(XL)
    return XL


# ------------------------------------------------------------------ run
def main():
    rng = random.Random(20260812)
    p_bo, p_gs, p_pf = 8.00, 6.00, 0.50      # non-zero panel fee, so it is really exercised
    rows = make_rows(TEST_ROWS, rng)
    payments = [(date(2026, 8, 5), "Olivia", 500.00),
                (date(2026, 8, 9), "Bob", 120.00),
                (date(2026, 8, 20), "Olivia", 75.00)]

    path = fill(rows, p_bo, p_gs, p_pf, payments)
    print(f"filled test workbook: {path}  ({TEST_ROWS} rows)")

    xl = formulas.ExcelModel().loads(path).finish()
    sol = xl.calculate()

    fname = path.split('/')[-1]

    def cell(sheet, ref):
        key = f"'[{fname}]{sheet.upper()}'!{ref}"
        return sol[key].value[0, 0]

    # ---- per-row -----------------------------------------------------
    exp = []
    for i, row in enumerate(rows):
        r = B.FIRST + i
        m = model_row(row["counts"], p_bo, p_gs, p_pf, row["other"])
        exp.append(m)
        for col in "HLMQS":
            check(f"Daily!{col}{r}", cell("Daily", f"{col}{r}"), m[col])
        check(f"Daily!R{r}", str(cell("Daily", f"R{r}")), direction_text(m["S"]))
        # prices must have flowed from the Rates tab
        check(f"Daily!I{r}", cell("Daily", f"I{r}"), Fr(str(p_bo)))
        check(f"Daily!J{r}", cell("Daily", f"J{r}"), Fr(str(p_gs)))
        check(f"Daily!K{r}", cell("Daily", f"K{r}"), Fr(str(p_pf)))
    print(f"  per-row cells checked: {TEST_ROWS * 9}")

    # ---- a blank row must stay blank, not read 0.00 -------------------
    blank = B.FIRST + TEST_ROWS - 1
    # (all TEST_ROWS are filled; check the row after the last one is not in range)
    # instead verify the guard directly on an emptied row
    from openpyxl import load_workbook as lw
    wb2 = lw(path)
    d2 = wb2["Daily"]
    for c in range(1, 20):
        d2.cell(row=blank, column=c).value = d2.cell(row=blank, column=c).value \
            if c not in (1, 2, 3, 4, 5, 6, 7, 14, 15, 16) else None
    blankpath = path.replace(".xlsx", "-blank.xlsx")
    wb2.save(blankpath)
    xl2 = formulas.ExcelModel().loads(blankpath).finish()
    sol2 = xl2.calculate()
    bn = blankpath.split('/')[-1]
    for col in "HIJKLMQRS":
        v = sol2[f"'[{bn}]DAILY'!{col}{blank}"].value[0, 0]
        if str(v) != "":
            FAILS.append(f"blank row Daily!{col}{blank} should be empty, got {v!r}")
    print("  blank-row guard checked on 9 columns")

    # ---- summary ------------------------------------------------------
    ws = load_workbook(path)["Summary"]
    def find(label):
        for r in range(1, ws.max_row + 1):
            if ws.cell(row=r, column=2).value == label:
                return r
        raise KeyError(label)

    tot_L = sum(m["L"] for m in exp)
    tot_M = sum(m["M"] for m in exp)
    tot_Q = sum(m["Q"] for m in exp)
    tot_S = sum(m["S"] for m in exp)

    check("Summary Olivia owes Bob",
          cell("Summary", f"C{find('Olivia owes Bob — lead sales')}"), tot_L)
    check("Summary Bob owes Olivia",
          cell("Summary", f"C{find('Bob owes Olivia — GS leads to his buyers, and panel fee')}"),
          tot_M)
    check("Summary other-cost adj",
          cell("Summary", f"C{find('Other cost adjustment')}"), tot_Q)
    net_r = find("NET FOR THE PERIOD")
    check("Summary NET", cell("Summary", f"C{net_r}"), tot_S)
    for r_ in range(1, ws.max_row + 1):
        lab = ws.cell(row=r_, column=2).value
        if isinstance(lab, str) and lab[:2] in ("1.", "2.", "3.", "4."):
            check(f"Summary check {lab[:2]} on clean data",
                  cell("Summary", f"C{r_}"), 0)

    # every lane total
    for label, idx in [("Bob's own campaign -> Olivia", 0), ("S -> Olivia", 1),
                       ("GS -> Olivia", 2),
                       ("Bob's own campaign -> Bob's private buyer", 3),
                       ("S -> Bob's private buyer", 4),
                       ("GS -> Bob's private buyer", 5)]:
        want = sum(r["counts"][idx] for r in rows)
        check(f"Summary lane '{label}'", cell("Summary", f"C{find(label)}"), want)

    paid_olivia = sum(a for _, w, a in payments if w == "Olivia")
    paid_bob = sum(a for _, w, a in payments if w == "Bob")
    check("Summary paid by Olivia", cell("Summary", f"C{find('Paid by Olivia to Bob')}"),
          Fr(str(paid_olivia)))
    check("Summary paid by Bob", cell("Summary", f"C{find('Paid by Bob to Olivia')}"),
          Fr(str(paid_bob)))
    out_r = find("STILL OUTSTANDING")
    want_out = tot_S - Fr(str(paid_olivia)) + Fr(str(paid_bob))
    check("Summary STILL OUTSTANDING", cell("Summary", f"C{out_r}"), want_out)

    sentence = str(cell("Summary", f"B{out_r + 2}"))
    ro = round(float(want_out), 2)
    want_sent = (f"OLIVIA PAYS BOB {float(want_out):,.2f}" if ro > 0 else
                 f"BOB PAYS OLIVIA {float(-want_out):,.2f}" if ro < 0 else
                 "NOTHING OWED — SETTLED")
    check("Summary sentence", sentence, want_sent)
    print("  summary: 15 figures checked")

    # ---- date window actually filters ---------------------------------
    wb3 = load_workbook(path)
    s3 = wb3["Summary"]
    s3["C4"] = date(2026, 8, 1)
    s3["C5"] = date(2026, 8, 10)          # first 10 rows only
    wpath = path.replace(".xlsx", "-window.xlsx")
    wb3.save(wpath)
    sol3 = formulas.ExcelModel().loads(wpath).finish().calculate()
    wn = wpath.split('/')[-1]
    win_net = sol3[f"'[{wn}]SUMMARY'!C{net_r}"].value[0, 0]
    check("Summary windowed NET (first 10 days)", win_net, sum(m["S"] for m in exp[:10]))
    # payments are windowed too: only the 5 Aug and 9 Aug ones fall inside
    win_out = sol3[f"'[{wn}]SUMMARY'!C{out_r}"].value[0, 0]
    check("Summary windowed OUTSTANDING",
          win_out, sum(m["S"] for m in exp[:10]) - Fr("500") + Fr("120"))
    print("  date window: 2 figures checked")

    # ---- the Example tab -----------------------------------------------
    real = "/root/workspace/lead-ledger-bob-olivia/lead-ledger-bob-olivia.xlsx"
    ex = [([40, 0, 25, 10, 0, 12], 8.0, 6.0, 0.0, None),
          ([0, 35, 18, 0, 8, 20], 8.0, 6.0, 0.0, (30.0, "Bob", "YES")),
          ([5, 0, 30, 2, 0, 45], 8.0, 6.0, 0.0, (50.0, "Olivia", "YES"))]
    solx = formulas.ExcelModel().loads(real).finish().calculate()
    rn = real.split('/')[-1]
    ex_net = []
    for i, (counts, a, b_, c_, other) in enumerate(ex):
        r = B.FIRST + i
        m = model_row(counts, a, b_, c_, other)
        ex_net.append(m["S"])
        for col in "HLMQS":
            check(f"Example!{col}{r}", solx[f"'[{rn}]EXAMPLE'!{col}{r}"].value[0, 0], m[col])
    print(f"  example tab: 15 cells checked   day nets = "
          f"{[float(x) for x in ex_net]}  period = {float(sum(ex_net))}")

    # ---- report ---------------------------------------------------------
    print()
    if FAILS:
        print(f"FAIL — {len(FAILS)} mismatch(es):")
        for f in FAILS[:40]:
            print("   ", f)
        sys.exit(1)
    print("PASS — every workbook formula matches the independent model.")


if __name__ == "__main__":
    main()
